import openpyxl
from openpyxl.styles import Font

# Create a new workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Simulation Verification"

# Define input parameters
inputs = {
    "Production Rate (kg/hour)": 100,
    "Operator Productivity": 0.85,
    "Actual Demand (kg)": 5000,
    "Shifts Per Day": 3,
    "Shift 1 Hours": 8,
    "Shift 2 Hours": 8,
    "Shift 3 Hours": 8,
}

# Write the input parameters
sheet["A1"] = "Input Parameters"
sheet["A1"].font = Font(bold=True)
row = 2
for key, value in inputs.items():
    sheet.cell(row=row, column=1).value = key
    sheet.cell(row=row, column=2).value = value
    row += 1

# Define column headers for simulation data
headers = [
    "Day", "Shift", "Start Time", "End Time", "Effective Hours", 
    "Downtime (min)", "Net Production (kg)", "Cumulative Production (kg)"
]

# Write headers for simulation data
start_row = row + 2
sheet.cell(row=start_row, column=1).value = "Simulation Data"
sheet.cell(row=start_row, column=1).font = Font(bold=True)
for col, header in enumerate(headers, start=1):
    sheet.cell(row=start_row + 1, column=col).value = header
    sheet.cell(row=start_row + 1, column=col).font = Font(bold=True)

# Define shift schedule
shifts = [
    {"shift": 1, "start_time": "00:00", "end_time": "08:00", "hours_cell": f"B4"},
    {"shift": 2, "start_time": "08:00", "end_time": "16:00", "hours_cell": f"B5"},
    {"shift": 3, "start_time": "16:00", "end_time": "23:59", "hours_cell": f"B6"},
]

# Initialize cumulative production
cumulative_production = 0
row = start_row + 2
day = 1

# Populate shift data
while cumulative_production < inputs["Actual Demand (kg)"]:
    for shift in shifts:
        if cumulative_production >= inputs["Actual Demand (kg)"]:
            break

        # Calculate effective hours and net production in Python
        effective_hours = shift["hours_cell"]
        downtime_cell = f"F{row}"  # Downtime
        net_production_formula = f"((B1 * ({effective_hours} * B2)) - ({downtime_cell}/60 * B1))"
        cumulative_production += inputs["Production Rate (kg/hour)"] * inputs["Operator Productivity"] * inputs[f"Shift {shift['shift']} Hours"]

        # Write data for each shift
        sheet.cell(row=row, column=1).value = day  # Day
        sheet.cell(row=row, column=2).value = shift["shift"]  # Shift
        sheet.cell(row=row, column=3).value = shift["start_time"]  # Start Time
        sheet.cell(row=row, column=4).value = shift["end_time"]  # End Time
        sheet.cell(row=row, column=5).value = effective_hours  # Effective Hours (formula)
        sheet.cell(row=row, column=6).value = 0  # Downtime (editable)
        sheet.cell(row=row, column=7).value = f"={net_production_formula}"  # Net Production (formula)
        sheet.cell(row=row, column=8).value = cumulative_production  # Python's cumulative production value

        row += 1
    day += 1

# Adjust column widths for readability
for col in range(1, 9):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

# Save the workbook
workbook.save("Simulation_Verification_Enhanced.xlsx")
print("Excel file created: Simulation_Verification_Enhanced.xlsx")